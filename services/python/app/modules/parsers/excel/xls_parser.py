import io

import xlrd
from openpyxl import Workbook


class XLSParser:
    """Parser for Microsoft Excel .xls files"""

    def __init__(self) -> None:
        pass

    def convert_xls_to_xlsx(self, binary: bytes) -> bytes:
        """
        Convert .xls file to .xlsx using xlrd and openpyxl (pure Python)

        Args:
            binary (bytes): The binary content of the XLS file

        Returns:
            bytes: The binary content of the converted XLSX file

        Raises:
            Exception: For conversion errors
        """
        try:
            # Read the .xls file using xlrd
            xls_workbook = xlrd.open_workbook(file_contents=binary)

            # Create a new .xlsx workbook using openpyxl
            xlsx_workbook = Workbook()
            # Remove the default sheet created by openpyxl
            xlsx_workbook.remove(xlsx_workbook.active)

            # Iterate through all sheets in the .xls file
            for sheet_name in xls_workbook.sheet_names():
                xls_sheet = xls_workbook.sheet_by_name(sheet_name)
                xlsx_sheet = xlsx_workbook.create_sheet(title=sheet_name)

                # Copy all cell values
                for row_idx in range(xls_sheet.nrows):
                    for col_idx in range(xls_sheet.ncols):
                        cell = xls_sheet.cell(row_idx, col_idx)
                        value = cell.value

                        # Handle different cell types
                        if cell.ctype == xlrd.XL_CELL_DATE:
                            # Convert Excel date to Python datetime
                            try:
                                value = xlrd.xldate_as_datetime(
                                    cell.value, xls_workbook.datemode
                                )
                            except Exception:
                                pass  # Keep original value if conversion fails

                        # openpyxl uses 1-based indexing
                        xlsx_sheet.cell(
                            row=row_idx + 1, column=col_idx + 1, value=value
                        )

            # Save to bytes
            output = io.BytesIO()
            xlsx_workbook.save(output)
            output.seek(0)

            return output.read()

        except Exception as e:
            raise Exception(f"Error converting .xls to .xlsx: {str(e)}") from e
